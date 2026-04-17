from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook

def save_xlsx_from_jsonl(input_path: str, output_path: str, columns: list[str], sheet_name: str) -> None:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name[:31] or "Ozon")
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    ws.append(columns)
    with open(input_path, "r", encoding="utf-8") as fh:
        for line in fh:
            if not line.strip():
                continue
            row = json.loads(line)
            ws.append([row.get(col, "") for col in columns])
    wb.save(output_path)


def export_worker_cli() -> int:
    try:
        input_path = sys.argv[2]
        output_path = sys.argv[3]
        columns = json.loads(sys.argv[4])
        sheet_name = sys.argv[5] if len(sys.argv) > 5 else "Ozon"
        save_xlsx_from_jsonl(input_path, output_path, columns, sheet_name)
        return 0
    except Exception as exc:
        err_path = sys.argv[3] + ".error.txt" if len(sys.argv) > 3 else "export.error.txt"
        try:
            Path(err_path).write_text(str(exc), encoding="utf-8")
        except Exception:
            pass
        return 1


